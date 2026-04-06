from flask import Flask, render_template, request, redirect, session, send_file
from database import obtenerusuarios, estudianteexiste, insertarestudiante, carguemasivo
from dashprincipal import creartablero
from config import SECRET_KEY, PORT, DEBUG
import pandas as pd
import io

app = Flask(__name__)
app.secret_key = SECRET_KEY

creartablero(app)


# ─── RUTA LOGIN ───────────────────────────────────────────────────────────────
@app.route("/", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        usuario  = obtenerusuarios(username)
        if usuario:
            if usuario["password"] == password:
                session["username"] = usuario["username"]
                session["rol"]      = usuario["rol"]
                return redirect("/dashprincipal/")
            else:
                error = "Contraseña incorrecta"
        else:
            error = "Usuario no existe"
    return render_template("login.html", error=error)


# ─── RUTA REGISTRAR ESTUDIANTE (acceso público) ───────────────────────────────
@app.route("/registrar", methods=["GET", "POST"])
def registrar():
    mensaje  = None
    tipo     = None
    carreras = ["Fisica", "Ingenieria", "Matematicas"]
    usuario_session = session.get("username", None)

    if request.method == "POST":
        nombre  = request.form.get("nombre",  "").strip()
        edad    = request.form.get("edad",    "").strip()
        carrera = request.form.get("carrera", "").strip()
        nota1   = request.form.get("nota1",   "").strip()
        nota2   = request.form.get("nota2",   "").strip()
        nota3   = request.form.get("nota3",   "").strip()

        if not all([nombre, edad, carrera, nota1, nota2, nota3]):
            mensaje = "Por favor completa todos los campos."
            tipo    = "error"
        elif not all(0 <= float(n) <= 5 for n in [nota1, nota2, nota3]):
            mensaje = "Las notas deben estar entre 0.0 y 5.0."
            tipo    = "error"
        elif estudianteexiste(nombre, carrera):
            mensaje = (f"⚠️ El estudiante '{nombre}' ya está registrado en "
                       f"{carrera}. No se puede duplicar el registro.")
            tipo    = "error"
        else:
            insertarestudiante(nombre, edad, carrera, nota1, nota2, nota3)
            mensaje = f"✅ Estudiante '{nombre}' registrado exitosamente en {carrera}."
            tipo    = "exito"

    return render_template("registrar.html",
                           mensaje=mensaje, tipo=tipo,
                           carreras=carreras, usuario=usuario_session)


# ─── RUTA CARGUE MASIVO ───────────────────────────────────────────────────────
@app.route("/carguemasivo", methods=["GET", "POST"])
def cargue_masivo():
    resultado       = None
    usuario_session = session.get("username", None)

    if request.method == "POST":
        archivo = request.files.get("archivo_excel")

        if not archivo or archivo.filename == "":
            resultado = {
                'tipo': 'error',
                'mensaje': "No se seleccionó ningún archivo.",
                'detalle': [],
                'hay_rechazados': False
            }
        elif not archivo.filename.endswith(('.xlsx', '.xls')):
            resultado = {
                'tipo': 'error',
                'mensaje': "El archivo debe ser Excel (.xlsx o .xls).",
                'detalle': [],
                'hay_rechazados': False
            }
        else:
            try:
                df_excel = pd.read_excel(archivo)
                resumen  = carguemasivo(df_excel)

                if resumen['insertados'] > 0:
                    session['cargue_masivo'] = True

                # Guardar rechazados en sesión para la descarga
                hay_rechazados = not resumen['rechazados'].empty
                if hay_rechazados:
                    session['rechazados'] = resumen['rechazados'].to_json(orient='records', force_ascii=False)

                partes = []
                if resumen['insertados'] > 0:
                    partes.append(f"✅ {resumen['insertados']} estudiante(s) insertado(s)")
                if resumen['duplicados'] > 0:
                    partes.append(f"⚠️ {resumen['duplicados']} duplicado(s) omitido(s)")
                if resumen['errores'] > 0:
                    partes.append(f"❌ {resumen['errores']} fila(s) con error")

                total_rechazados = resumen['duplicados'] + resumen['errores']

                resultado = {
                    'tipo':           'exito' if resumen['insertados'] > 0 else 'error',
                    'mensaje':        " — ".join(partes) if partes else "No se procesó ningún registro.",
                    'detalle':        resumen['detalle_errores'],
                    'hay_rechazados': hay_rechazados,
                    'stats': {
                        'insertados':  resumen['insertados'],
                        'rechazados':  total_rechazados,
                        'duplicados':  resumen['duplicados'],
                        'total':       resumen['insertados'] + total_rechazados
                    }
                }

            except Exception as e:
                resultado = {
                    'tipo': 'error',
                    'mensaje': f"Error al leer el archivo: {str(e)}",
                    'detalle': [],
                    'hay_rechazados': False
                }

    return render_template("carguemasivo.html",
                           resultado=resultado,
                           usuario=usuario_session)


# ─── RUTA DESCARGA RECHAZADOS ─────────────────────────────────────────────────
@app.route("/descargar_rechazados")
def descargar_rechazados():
    from openpyxl.styles import PatternFill, Font

    rechazados_json = session.get('rechazados', None)

    if not rechazados_json:
        return redirect("/carguemasivo")

    df = pd.read_json(io.StringIO(rechazados_json), orient='records')

    # Generar Excel en memoria
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Rechazados')

        ws        = writer.sheets['Rechazados']
        red_fill  = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        bold_font = Font(bold=True)

        # Formato encabezados
        for cell in ws[1]:
            cell.font = bold_font
            if cell.value and "Motivo" in str(cell.value):
                cell.fill = red_fill

        # Ajustar ancho de columnas
        for col in ws.columns:
            max_len = max((len(str(cell.value)) if cell.value is not None else 0) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 55)

    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='estudiantes_rechazados.xlsx'
    )


# ─── RUTA LOGOUT ──────────────────────────────────────────────────────────────
@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")


if __name__ == "__main__":
    port = int(os.environ.get("PORT", PORT))
    app.run(debug=DEBUG, port=port, host="0.0.0.0")
